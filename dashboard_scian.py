"""
Dashboard SCIAN — Streamlit
Ejecutar:
  & "C:\\ruta\\python.exe" -m streamlit run "C:\\ruta\\dashboard_scian.py"
"""

import io, json, os, re
from collections import defaultdict, Counter
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="SCIAN Analytics", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    .stApp { background-color: #0F1117; }
    section[data-testid="stSidebar"] { background-color: #161B22; }
    .sec-title {
        font-size:1.05rem; font-weight:600; color:#E2E8F0;
        border-left:3px solid #2E75B6; padding-left:12px;
        margin:28px 0 16px 0; letter-spacing:0.03em;
    }
    hr.div { border:none; border-top:1px solid #2A3441; margin:24px 0; }
    #MainMenu, footer { visibility:hidden; }
</style>
""", unsafe_allow_html=True)

GRID = dict(gridcolor="#2A3441", linecolor="#2A3441")
BASE = dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#CBD5E1", family="Inter, sans-serif"),
            legend=dict(bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=20, r=20, t=40, b=20))

def lay(fig, **kw):
    fig.update_layout(**{**BASE, **kw}); return fig

def pch(fig): st.plotly_chart(fig, width="stretch")

def kpi(label, value, sub=""):
    sub_html = f"<p style='font-size:.85rem;color:#CBD5E1;margin-top:4px;'>{sub}</p>" if sub else ""
    st.markdown(
        "<div style='background:linear-gradient(135deg,#161B22,#1F2937);"
        "border:1px solid #2E75B6;border-radius:12px;padding:20px 24px;text-align:center;"
        "box-shadow:0 4px 20px rgba(46,117,182,.15);'>"
        f"<p style='font-size:2.2rem;font-weight:700;color:#60A5FA;line-height:1.1;margin:0;'>{value}</p>"
        f"<p style='font-size:.75rem;color:#94A3B8;text-transform:uppercase;"
        f"letter-spacing:.08em;margin-top:4px;'>{label}</p>{sub_html}</div>",
        unsafe_allow_html=True)

def sec(txt): st.markdown(f'<p class="sec-title">{txt}</p>', unsafe_allow_html=True)
def div():    st.markdown('<hr class="div">', unsafe_allow_html=True)

def bar_h(df_in, x_col, y_col, nom_col, pct_col, title):
    df_in = df_in.copy(); df_in[y_col] = df_in[y_col].astype(str)
    fig = go.Figure(go.Bar(
        x=df_in[x_col], y=df_in[y_col], orientation="h",
        marker=dict(color=df_in[x_col],
                    colorscale=[[0,"#1F4E79"],[1,"#60A5FA"]],
                    cmin=0, cmax=df_in[x_col].max()),
        customdata=list(zip(df_in[nom_col], df_in[pct_col])),
        hovertemplate="<b>%{y}</b><br>%{customdata[0]}<br>RFCs: %{x}<br>%{customdata[1]}%<extra></extra>",
        text=df_in[pct_col].astype(str)+"%", textposition="outside",
        textfont=dict(color="#94A3B8", size=10),
    ))
    lay(fig, title=title, xaxis=dict(title="RFCs", **GRID),
        yaxis=dict(type="category", categoryorder="total ascending",
                   gridcolor="#2A3441", tickfont=dict(size=11)))
    return fig

def grafica_combos(combos_df, title):
    combos_df = combos_df.copy()
    combos_df["combo"]   = combos_df["ppal_cod"] + " → " + combos_df["sec_cod"]
    combos_df["detalle"] = combos_df["ppal_nom"] + " → " + combos_df["sec_nom"]
    fig = go.Figure(go.Bar(
        x=combos_df["RFCs"], y=combos_df["combo"], orientation="h",
        marker_color="#2E75B6", customdata=combos_df["detalle"],
        text=combos_df["RFCs"], textposition="outside",
        textfont=dict(color="#94A3B8", size=10),
        hovertemplate="<b>%{y}</b><br>%{customdata}<br>RFCs: %{x}<extra></extra>",
    ))
    lay(fig, title=title, xaxis=dict(title="RFCs", **GRID),
        yaxis=dict(type="category", categoryorder="total ascending",
                   gridcolor="#2A3441", tickfont=dict(size=10)))
    return fig

def grafica_dist(dist_df, total, title):
    dist_df = dist_df.copy()
    dist_df["pct"] = (dist_df["RFCs"] / total * 100).round(1)
    fig = go.Figure(go.Bar(
        x=dist_df["N actividades"].astype(str), y=dist_df["RFCs"],
        marker_color="#2E75B6", customdata=dist_df["pct"],
        text=dist_df["pct"].astype(str)+"%", textposition="outside",
        textfont=dict(color="#94A3B8", size=10),
        hovertemplate="<b>%{x} actividades</b><br>RFCs: %{y}<br>%{customdata}%<extra></extra>",
    ))
    lay(fig, title=title,
        xaxis=dict(title="N° actividades", type="category", **GRID),
        yaxis=dict(title="RFCs", gridcolor="#2A3441"))
    return fig

def grafica_top_div(top_df, title):
    fig = go.Figure(go.Bar(
        x=top_df["Total Act."], y=top_df["RFC"], orientation="h",
        marker=dict(color=top_df["Total Act."],
                    colorscale=[[0,"#1F4E79"],[1,"#60A5FA"]]),
        customdata=list(zip(top_df["Tipo"], top_df["Secundarias"],
                            top_df.get("Company", [""] * len(top_df)))),
        text=top_df["Total Act."], textposition="outside",
        textfont=dict(color="#94A3B8", size=10),
        hovertemplate="<b>%{y}</b><br>%{customdata[0]}<br>"
                      "Company: %{customdata[2]}<br>"
                      "Total: %{x}<br>Secundarias: %{customdata[1]}<extra></extra>",
    ))
    lay(fig, title=title, xaxis=dict(title="N° actividades", **GRID),
        yaxis=dict(type="category", categoryorder="total ascending",
                   gridcolor="#2A3441", tickfont=dict(size=9)))
    return fig

def get_combos(df_ppal_in, df_sec_in, top=15):
    pm = df_ppal_in[["rfc","sub_cod","sub_nom"]].rename(
        columns={"sub_cod":"ppal_cod","sub_nom":"ppal_nom"})
    sm = df_sec_in[["rfc","sub_cod","sub_nom"]].rename(
        columns={"sub_cod":"sec_cod","sub_nom":"sec_nom"})
    return (pm.merge(sm, on="rfc")
            .groupby(["ppal_cod","ppal_nom","sec_cod","sec_nom"])
            .size().reset_index(name="RFCs")
            .sort_values("RFCs", ascending=False).head(top))

def get_div(df_in):
    return (df_in.groupby(["rfc","company","tipo"])
            .agg(total_act=("sub_cod","count"),
                 act_sec=("tipo_act", lambda x: (x=="Secundaria").sum()))
            .reset_index())


# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

CARPETA = os.path.dirname(os.path.abspath(__file__))

def expandir(cod_raw):
    cod = str(cod_raw).strip()
    m = re.match(r'^(\d+)-(\d+)$', cod)
    return [str(n) for n in range(int(m.group(1)), int(m.group(2))+1)] if m else [cod]

@st.cache_data
def cargar_catalogo():
    ruta = os.path.join(CARPETA, "scian_2023_categorias_y_productos.xlsx")
    if not os.path.exists(ruta): return {}, {}
    try:
        ds = pd.read_excel(ruta, sheet_name="SECTOR",    usecols=[0,1], header=0, dtype=str)
        db = pd.read_excel(ruta, sheet_name="SUBSECTOR", usecols=[0,1], header=0, dtype=str)
    except: return {}, {}
    def td(df):
        d, cols = {}, df.columns.tolist()
        for _, r in df.iterrows():
            c = str(r[cols[0]] or "").strip(); n = str(r[cols[1]] or "").strip()
            if c and n and c.lower() not in ("nan","none"):
                for x in expandir(c): d[x] = n
        return d
    return td(ds), td(db)

@st.cache_data
def cargar_json():
    ruta = os.path.join(CARPETA, "datos_scian.json")
    if not os.path.exists(ruta): return []
    with open(ruta, encoding="utf-8") as f:
        return json.load(f).get("todos_los_rfcs", [])

def tp(rfc):
    n = len(str(rfc).strip())
    if n == 12: return "Persona Moral"
    if n == 13: return "Persona Física"
    return "Desconocido"

@st.cache_data
def build_df(rfcs_json, cat_sec, cat_sub):
    rows = []
    for e in json.loads(rfcs_json):
        rfc  = str(e.get("Rfc","")).strip()
        cid  = str(e.get("CompanyId","Sin Company"))
        tipo = tp(rfc)
        subs = sorted(e.get("ScianSubsector",[]),
                      key=lambda x: x.get("Percentage",0), reverse=True)
        for i, s in enumerate(subs):
            cod = str(s.get("CodigoSubsectorSCIAN") or "N/A").strip()
            sc  = cod[:2] if len(cod) >= 2 else "N/A"
            rows.append({"rfc": rfc, "company": cid, "tipo": tipo,
                         "tipo_act": "Principal" if i==0 else "Secundaria",
                         "sub_cod": cod,
                         "sub_nom": cat_sec.get(cod) or str(s.get("NameSubsectorSCIAN") or "Sin nombre"),
                         "sec_cod": sc,
                         "sec_nom": cat_sec.get(sc, "Sin clasificar"),
                         "pct": s.get("Percentage",0)})
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# GENERADOR DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def generar_excel(rfcs_json, cat_sec_items, cat_sub_items):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cat_sec = dict(cat_sec_items)
    cat_sub = dict(cat_sub_items)
    rfcs    = json.loads(rfcs_json)

    C_AZUL = "1F4E79"; C_MED = "2E75B6"; C_CLAR = "D6E4F0"
    C_VERDE = "D9EAD3"; C_AMAR = "FFF2CC"

    def borde():
        l = Side(style="thin", color="BFBFBF")
        return Border(left=l, right=l, top=l, bottom=l)

    def enc(cell, bg=C_AZUL, fg="FFFFFF", sz=11, center=True):
        cell.font      = Font(bold=True, color=fg, size=sz)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center", wrap_text=True)
        cell.border    = borde()

    def cel(cell, bold=False, center=False, bg=None, sz=10):
        cell.font      = Font(bold=bold, size=sz)
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center")
        cell.border    = borde()
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    def ajustar(ws):
        for col in ws.columns:
            mx = 0; l = get_column_letter(col[0].column)
            for c in col:
                try: mx = max(mx, len(str(c.value or "")))
                except: pass
            ws.column_dimensions[l].width = min(mx + 4, 60)

    # Procesar RFCs
    registros = []
    rfcs_por_company = {}
    for e in rfcs:
        rfc  = str(e.get("Rfc","")).strip()
        cid  = str(e.get("CompanyId","Sin Company"))
        tipo = tp(rfc)
        subs = sorted(e.get("ScianSubsector",[]),
                      key=lambda x: x.get("Percentage",0), reverse=True)
        ppal = None; secs = []
        for i, s in enumerate(subs):
            cod = str(s.get("CodigoSubsectorSCIAN") or "N/A").strip()
            sc  = cod[:2] if len(cod) >= 2 else "N/A"
            info = {"cod": cod,
                    "nom": cat_sub.get(cod) or str(s.get("NameSubsectorSCIAN") or ""),
                    "sec_cod": sc, "sec_nom": cat_sec.get(sc,"Sin clasificar"),
                    "pct": s.get("Percentage",0)}
            if i == 0: ppal = info
            else: secs.append(info)
        registros.append({"rfc":rfc,"company":cid,"tipo":tipo,
                          "principal":ppal,"secundarias":secs})
        rfcs_por_company.setdefault(cid, []).append(
            {"rfc":rfc,"tipo":tipo,"principal":ppal,"secundarias":secs})

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # Hoja 1: Detalle por RFC
    ws = wb.create_sheet("1. Detalle por RFC")
    ws.freeze_panes = "A3"
    cols = ["Company ID","RFC","Tipo Persona","Tipo Actividad",
            "Cód. Sector","Nombre Sector","Cód. Subsector","Nombre Subsector","% Actividad"]
    for i,t in enumerate(cols,1): enc(ws.cell(row=2,column=i,value=t))
    fila = 3
    for cid in sorted(rfcs_por_company):
        ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=9)
        c = ws.cell(row=fila,column=1,value=f"  Company {cid} — {len(rfcs_por_company[cid])} RFCs")
        c.font=Font(bold=True,size=11,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=C_MED)
        c.alignment=Alignment(horizontal="left",vertical="center")
        c.border=borde(); fila+=1
        for reg in rfcs_por_company[cid]:
            p = reg["principal"]
            if p:
                for j,v in enumerate([cid,reg["rfc"],reg["tipo"],"PRINCIPAL",
                                       p["sec_cod"],p["sec_nom"],p["cod"],p["nom"],p["pct"]],1):
                    cel(ws.cell(row=fila,column=j,value=v),bold=True,
                        center=(j in [1,3,4,5,7,9]),bg=C_CLAR)
                fila+=1
            for s in reg["secundarias"]:
                for j,v in enumerate([cid,reg["rfc"],reg["tipo"],"secundaria",
                                       s["sec_cod"],s["sec_nom"],s["cod"],s["nom"],s["pct"]],1):
                    cel(ws.cell(row=fila,column=j,value=v),center=(j in [1,3,4,5,7,9]))
                fila+=1
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=9)
    c=ws.cell(row=fila,column=1,value=f"TOTAL RFCs: {len(rfcs)}")
    c.font=Font(bold=True,size=11); c.fill=PatternFill("solid",fgColor=C_CLAR)
    c.alignment=Alignment(horizontal="center"); c.border=borde()
    ajustar(ws)

    # Hoja 2: Resumen por Company
    ws2 = wb.create_sheet("2. Resumen por Company")
    cols2=["Company ID","Total RFCs","Persona Moral","Persona Física",
           "Solo Principal","Con Secundarias","Subsector Principal más frecuente"]
    for i,t in enumerate(cols2,1): enc(ws2.cell(row=1,column=i,value=t))
    for fila2,(cid,regs) in enumerate(sorted(rfcs_por_company.items()),2):
        nm=nf=solo=con=0; cnt=Counter()
        for r in regs:
            if "Moral"  in r["tipo"]: nm+=1
            if "Física" in r["tipo"]: nf+=1
            if not r["secundarias"]: solo+=1
            else: con+=1
            if r["principal"]: cnt[r["principal"]["cod"]+" — "+r["principal"]["nom"]]+=1
        top_sub = cnt.most_common(1)[0][0] if cnt else "—"
        for j,v in enumerate([cid,len(regs),nm,nf,solo,con,top_sub],1):
            cel(ws2.cell(row=fila2,column=j,value=v),center=(j!=7))
    ajustar(ws2)

    # Hoja 3: Global Subsectores Principales
    ws3 = wb.create_sheet("3. Global - Subsectores Ppal")
    cols3=["Cód. Sector","Nombre Sector","Cód. Subsector","Nombre Subsector",
           "Total RFCs","% del Total","Morales","Físicas"]
    for i,t in enumerate(cols3,1): enc(ws3.cell(row=1,column=i,value=t))
    sub_agg=defaultdict(lambda:{"nom":"","sc":"","sc_nom":"","rfcs":0,"m":0,"f":0})
    for reg in registros:
        p=reg["principal"]
        if not p: continue
        sub_agg[p["cod"]]["nom"]=p["nom"]; sub_agg[p["cod"]]["sc"]=p["sec_cod"]
        sub_agg[p["cod"]]["sc_nom"]=p["sec_nom"]; sub_agg[p["cod"]]["rfcs"]+=1
        if "Moral"  in reg["tipo"]: sub_agg[p["cod"]]["m"]+=1
        if "Física" in reg["tipo"]: sub_agg[p["cod"]]["f"]+=1
    total=len(registros)
    for fila3,(cod,d) in enumerate(sorted(sub_agg.items(),key=lambda x:-x[1]["rfcs"]),2):
        pct=round(d["rfcs"]/total*100,2) if total else 0
        for j,v in enumerate([d["sc"],d["sc_nom"],cod,d["nom"],d["rfcs"],pct,d["m"],d["f"]],1):
            cel(ws3.cell(row=fila3,column=j,value=v),center=(j not in [2,4]))
    ajustar(ws3)

    # Hoja 4: RFCs Duplicados
    ws4 = wb.create_sheet("4. RFCs Duplicados")
    rfc_comp=defaultdict(list)
    for reg in registros: rfc_comp[reg["rfc"]].append(reg["company"])
    dups={r:cs for r,cs in rfc_comp.items() if len(cs)>1}
    if not dups:
        ws4.merge_cells("A1:D1")
        c=ws4.cell(row=1,column=1,value="✅  No se encontraron RFCs duplicados entre companies.")
        c.font=Font(bold=True,size=11,color="375623")
        c.fill=PatternFill("solid",fgColor=C_VERDE)
        c.alignment=Alignment(horizontal="center"); c.border=borde()
    else:
        for i,t in enumerate(["RFC","Tipo Persona","# Companies","Companies"],1):
            enc(ws4.cell(row=1,column=i,value=t))
        for fila4,(rfc,cs) in enumerate(sorted(dups.items(),key=lambda x:-len(x[1])),2):
            for j,v in enumerate([rfc,tp(rfc),len(cs)," | ".join(str(c) for c in cs)],1):
                cel(ws4.cell(row=fila4,column=j,value=v),center=(j in [2,3]),bg=C_AMAR)
    ajustar(ws4)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# INICIALIZACIÓN
# ══════════════════════════════════════════════════════════════════════════════

cat_sec, cat_sub = cargar_catalogo()
rfcs_raw         = cargar_json()

if not rfcs_raw:
    st.error("No se encontró `datos_scian.json`."); st.stop()

df   = build_df(json.dumps(rfcs_raw), cat_sec, cat_sub)
df_p = df[df["tipo_act"] == "Principal"].copy()
df_s = df[df["tipo_act"] == "Secundaria"].copy()
TOT     = df_p["rfc"].nunique()
TOT_CMP = df_p["company"].nunique()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("<h2 style='color:#60A5FA;font-size:1.3rem;margin-bottom:4px;'>📊 SCIAN Analytics</h2>"
                "<p style='color:#64748B;font-size:.8rem;margin-top:0;'>Sistema de Clasificación Industrial</p>",
                unsafe_allow_html=True)
    st.divider()
    vista = st.radio("Vista", ["🌐  Global", "🏢  Por Empresa"], label_visibility="collapsed")
    st.divider()
    st.markdown("<p style='color:#94A3B8;font-size:.8rem;margin-bottom:6px;'>📥 Exportar reporte</p>",
                unsafe_allow_html=True)
    excel_bytes = generar_excel(
        json.dumps(rfcs_raw),
        tuple(cat_sec.items()),
        tuple(cat_sub.items()),
    )
    st.download_button(
        label="⬇️  Descargar Excel",
        data=excel_bytes,
        file_name="reporte_scian.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.divider()
    st.markdown("<p style='color:#475569;font-size:.75rem;'>Los datos se actualizan "
                "al reemplazar <code>datos_scian.json</code>.</p>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# VISTA GLOBAL
# ══════════════════════════════════════════════════════════════════════════════

if vista == "🌐  Global":

    st.markdown("<h1 style='color:#E2E8F0;font-size:1.8rem;font-weight:700;margin-bottom:4px;'>"
                "Vista Global</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#64748B;font-size:.9rem;margin-top:0;'>"
                "Universo total de RFCs · actividad principal por RFC</p>", unsafe_allow_html=True)
    div()

    # KPIs
    n_m = (df_p["tipo"]=="Persona Moral").sum()
    n_f = (df_p["tipo"]=="Persona Física").sum()
    dup = df_p.groupby("rfc")["company"].nunique()
    n_d = (dup > 1).sum()

    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1: kpi("Total RFCs",    f"{TOT:,}")
    with c2: kpi("Companies",      f"{TOT_CMP:,}")
    with c3: kpi("Sectores",       f"{df_p['sec_cod'].nunique():,}")
    with c4: kpi("Subsectores",    f"{df_p['sub_cod'].nunique():,}")
    with c5: kpi("Persona Moral",  f"{n_m:,}", f"{round(n_m/TOT*100,1)}% del total")
    with c6: kpi("Persona Física", f"{n_f:,}", f"{round(n_f/TOT*100,1)}% del total")
    div()

    # Composición tipo persona
    sec("Composición por Tipo de Persona")
    ga, gb = st.columns([1,2])
    with ga:
        dt = df_p["tipo"].value_counts().reset_index(); dt.columns = ["Tipo","Total"]
        fig = px.pie(dt, values="Total", names="Tipo", hole=0.62,
                     color_discrete_sequence=["#2E75B6","#F59E0B","#64748B"])
        fig.update_traces(textposition="outside", textinfo="percent+label",
                          textfont_color="#CBD5E1")
        lay(fig, showlegend=False, title=""); pch(fig)
    with gb:
        dts = df_p.groupby(["sec_cod","tipo"]).size().reset_index(name="n")
        dts["sec_label"] = dts["sec_cod"].map(lambda c: cat_sec.get(c, c))
        fig = px.bar(dts, x="n", y="sec_cod", color="tipo", orientation="h",
                     barmode="stack", labels={"n":"RFCs","sec_cod":"","tipo":""},
                     color_discrete_map={"Persona Moral":"#2E75B6",
                                         "Persona Física":"#F59E0B",
                                         "Desconocido":"#64748B"},
                     title="Moral vs Física por Sector", custom_data=["sec_label"])
        fig.update_traces(
            hovertemplate="<b>%{y}</b> · %{customdata[0]}<br>RFCs: %{x}<extra></extra>")
        lay(fig, xaxis=dict(title="RFCs",**GRID),
            yaxis=dict(type="category", gridcolor="#2A3441", tickfont=dict(size=11)))
        pch(fig)
    div()

    # Actividad principal — sectores
    sec("Actividad Principal — Sectores")
    ds = (df_p.groupby(["sec_cod","sec_nom"]).size()
          .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
    ds["pct"] = (ds["RFCs"]/TOT*100).round(1)
    ga, gb = st.columns(2)
    with ga:
        pch(bar_h(ds.head(15), "RFCs","sec_cod","sec_nom","pct","Top 15 Sectores — Principal"))
    with gb:
        ds["label"] = ds["sec_cod"] + " · " + ds["sec_nom"]
        fig = px.treemap(ds, path=["label"], values="RFCs",
                         title="Proporción de Sectores — Principal", color="RFCs",
                         color_continuous_scale=["#1F4E79","#60A5FA"],
                         custom_data=["sec_cod","pct"])
        fig.update_traces(
            texttemplate="<b>%{customdata[0]}</b>",
            hovertemplate="<b>%{label}</b><br>RFCs: %{value}<br>%{customdata[1]}%<extra></extra>",
            textfont_color="white")
        lay(fig, coloraxis_showscale=False); pch(fig)
    div()

    # Actividad principal — subsectores
    sec("Actividad Principal — Subsectores")
    db = (df_p.groupby(["sub_cod","sub_nom"]).size()
          .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
    db["pct"]  = (db["RFCs"]/TOT*100).round(1)
    db["acum"] = db["pct"].cumsum().round(1)
    top_n = st.slider("Top N subsectores (principal)", 5, min(50,len(db)), 15, key="sl_ppal")
    dt    = db.head(top_n).copy().reset_index(drop=True)
    cats  = dt["sub_cod"].astype(str).tolist()
    ga, gb = st.columns(2)
    with ga:
        pch(bar_h(dt, "RFCs","sub_cod","sub_nom","pct", f"Top {top_n} Subsectores — Principal"))
    with gb:
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=cats, y=dt["RFCs"], name="RFCs", marker_color="#2E75B6",
            customdata=list(zip(dt["sub_nom"], dt["pct"])),
            hovertemplate="<b>%{x}</b><br>%{customdata[0]}<br>RFCs: %{y} · %{customdata[1]}%<extra></extra>",
        ))
        fig.add_trace(go.Scatter(
            x=cats, y=dt["acum"], name="% Acumulado", yaxis="y2",
            line=dict(color="#F59E0B", width=2), mode="lines+markers", marker=dict(size=5),
            hovertemplate="%{y:.1f}% acumulado<extra></extra>",
        ))
        fig.add_hline(y=80, line_dash="dot", line_color="#EF4444", yref="y2",
                      annotation_text="80%", annotation_font_color="#EF4444")
        lay(fig, title="Curva de Pareto — Subsectores Principal",
            xaxis=dict(type="category", tickangle=-45, gridcolor="#2A3441", tickfont=dict(size=10)),
            yaxis=dict(gridcolor="#2A3441"),
            yaxis2=dict(overlaying="y", side="right", range=[0,105],
                        ticksuffix="%", gridcolor="rgba(0,0,0,0)"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            barmode="overlay")
        pch(fig)
    div()

    # Actividad secundaria — sectores
    if not df_s.empty:
        sec("Actividad Secundaria — Sectores")
        ds2 = (df_s.groupby(["sec_cod","sec_nom"])["rfc"].nunique()
               .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
        ds2["pct"] = (ds2["RFCs"]/TOT*100).round(1)
        ga, gb = st.columns(2)
        with ga:
            pch(bar_h(ds2.head(15), "RFCs","sec_cod","sec_nom","pct","Top 15 Sectores — Secundaria"))
        with gb:
            ds2["label"] = ds2["sec_cod"] + " · " + ds2["sec_nom"]
            fig = px.treemap(ds2, path=["label"], values="RFCs",
                             title="Proporción de Sectores — Secundaria", color="RFCs",
                             color_continuous_scale=["#375623","#6AAB49"],
                             custom_data=["sec_cod","pct"])
            fig.update_traces(
                texttemplate="<b>%{customdata[0]}</b>",
                hovertemplate="<b>%{label}</b><br>RFCs: %{value}<br>%{customdata[1]}%<extra></extra>",
                textfont_color="white")
            lay(fig, coloraxis_showscale=False); pch(fig)
        div()

        sec("Actividad Secundaria — Subsectores")
        db2 = (df_s.groupby(["sub_cod","sub_nom"])["rfc"].nunique()
               .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
        db2["pct"] = (db2["RFCs"]/TOT*100).round(1)
        top_n2 = st.slider("Top N subsectores (secundaria)", 5, min(50,len(db2)), 15, key="sl_sec")
        pch(bar_h(db2.head(top_n2), "RFCs","sub_cod","sub_nom","pct",
                  f"Top {top_n2} Subsectores — Secundaria"))
        div()

    # Diversificación
    sec("Diversificación — RFCs con Actividades Secundarias")
    div_df    = get_div(df)
    solo_ppal = (div_df["act_sec"]==0).sum()
    con_sec   = (div_df["act_sec"]>0).sum()
    prom_act  = round(div_df["total_act"].mean(), 1)
    max_act   = div_df["total_act"].max()

    d1,d2,d3,d4 = st.columns(4)
    with d1: kpi("Solo actividad principal",    f"{solo_ppal:,}", f"{round(solo_ppal/TOT*100,1)}%")
    with d2: kpi("Con actividades secundarias", f"{con_sec:,}",   f"{round(con_sec/TOT*100,1)}%")
    with d3: kpi("Promedio de actividades",     f"{prom_act}")
    with d4: kpi("Máx. actividades en un RFC",  f"{max_act}")
    st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)

    ga, gb = st.columns(2)
    with ga:
        dist = (div_df["total_act"].value_counts()
                .reset_index().rename(columns={"total_act":"N actividades","count":"RFCs"})
                .sort_values("N actividades"))
        pch(grafica_dist(dist, TOT, "Distribución: N° de actividades por RFC"))
    with gb:
        top_div = (div_df[div_df["act_sec"]>0]
                   .sort_values("total_act", ascending=False).head(15)
                   [["rfc","company","tipo","total_act","act_sec"]]
                   .rename(columns={"rfc":"RFC","company":"Company","tipo":"Tipo",
                                    "total_act":"Total Act.","act_sec":"Secundarias"}))
        if not top_div.empty:
            pch(grafica_top_div(top_div, "Top 15 RFCs más diversificados"))

    combos = get_combos(df_p, df_s)
    if not combos.empty:
        pch(grafica_combos(combos, "Top 15 Combinaciones Principal → Secundaria"))
    div()

    if n_d > 0:
        sec(f"⚠️ RFCs en más de una Company — {n_d} encontrados")
        dd = dup[dup>1].reset_index().rename(columns={"company":"# Companies"})
        dd["Tipo"] = dd["rfc"].apply(tp)
        dd["Companies"] = dd["rfc"].apply(
            lambda r: " · ".join(sorted(df_p[df_p["rfc"]==r]["company"].unique())))
        st.dataframe(dd, width="stretch", hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# VISTA POR EMPRESA
# ══════════════════════════════════════════════════════════════════════════════

else:
    st.markdown("<h1 style='color:#E2E8F0;font-size:1.8rem;font-weight:700;margin-bottom:4px;'>"
                "Vista por Empresa</h1>", unsafe_allow_html=True)

    companies = sorted(df_p["company"].unique())

    with st.sidebar:
        emp_sel = st.selectbox("Seleccionar Company", options=companies,
                               format_func=lambda x: f"Company {x}")
        st.divider()
        comparar = st.checkbox("Comparar con otra company")
        emp_cmp  = None
        if comparar:
            otras = [c for c in companies if c != emp_sel]
            if otras:
                emp_cmp = st.selectbox("Company a comparar", options=otras,
                                       format_func=lambda x: f"Company {x}")

    df_e   = df_p[df_p["company"] == emp_sel]
    df_e_s = df_s[df_s["company"] == emp_sel]
    df_e_a = df[df["company"]  == emp_sel]

    st.markdown(
        f"<p style='color:#64748B;font-size:.9rem;margin-top:0;'>"
        f"<span style='background:#2E75B6;color:white;border-radius:6px;"
        f"padding:2px 10px;font-size:.8rem;font-weight:600;'>"
        f"Company {emp_sel}</span> Perfil sectorial detallado</p>",
        unsafe_allow_html=True)
    div()

    tr   = df_e["rfc"].nunique()
    ts   = df_e["sec_cod"].nunique()
    tb   = df_e["sub_cod"].nunique()
    nm   = (df_e["tipo"]=="Persona Moral").sum()
    nf   = (df_e["tipo"]=="Persona Física").sum()
    pem  = round(tr/TOT*100,1) if TOT else 0
    pm_e = round(nm/tr*100,1)  if tr  else 0
    pf_e = round(nf/tr*100,1)  if tr  else 0

    e1,e2,e3,e4,e5 = st.columns(5)
    with e1: kpi("RFCs",          f"{tr:,}",  f"{pem}% del global")
    with e2: kpi("Sectores",       f"{ts:,}")
    with e3: kpi("Subsectores",    f"{tb:,}")
    with e4: kpi("Persona Moral",  f"{nm:,}",  f"{pm_e}%")
    with e5: kpi("Persona Física", f"{nf:,}",  f"{pf_e}%")
    div()

    # Actividad principal — sectores
    sec("Actividad Principal — Sectores")
    dse = (df_e.groupby(["sec_cod","sec_nom"]).size()
           .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
    dse["pct"] = (dse["RFCs"]/tr*100).round(1)
    ea, eb = st.columns(2)
    with ea:
        pch(bar_h(dse, "RFCs","sec_cod","sec_nom","pct","RFCs por Sector — Principal"))
    with eb:
        fig = px.pie(dse, values="RFCs", names="sec_cod",
                     hole=0.55, title="Proporción por Sector — Principal",
                     custom_data=["sec_nom","pct"])
        fig.update_traces(
            textposition="outside", textinfo="percent+label", textfont_color="#CBD5E1",
            hovertemplate="<b>%{label}</b><br>%{customdata[0]}<br>"
                          "RFCs: %{value}<br>%{customdata[1]}%<extra></extra>")
        lay(fig); pch(fig)
    div()

    # Actividad principal — subsectores
    sec("Actividad Principal — Subsectores")
    dsu = (df_e.groupby(["sub_cod","sub_nom"]).size()
           .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
    dsu["pct"] = (dsu["RFCs"]/tr*100).round(1)
    top_ne = st.slider("Top N subsectores — principal", 5, min(30,len(dsu)), 10, key="sl_e")
    pch(bar_h(dsu.head(top_ne), "RFCs","sub_cod","sub_nom","pct",
              f"Top {top_ne} Subsectores Principales · Company {emp_sel}"))
    div()

    # Actividad secundaria
    if not df_e_s.empty:
        sec("Actividad Secundaria — Sectores y Subsectores")
        ds2e = (df_e_s.groupby(["sec_cod","sec_nom"])["rfc"].nunique()
                .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
        ds2e["pct"] = (ds2e["RFCs"]/tr*100).round(1)
        db2e = (df_e_s.groupby(["sub_cod","sub_nom"])["rfc"].nunique()
                .reset_index(name="RFCs").sort_values("RFCs", ascending=False))
        db2e["pct"] = (db2e["RFCs"]/tr*100).round(1)
        ea2, eb2 = st.columns(2)
        with ea2:
            pch(bar_h(ds2e, "RFCs","sec_cod","sec_nom","pct","Sectores — Actividad Secundaria"))
        with eb2:
            top_se2 = st.slider("Top N subsectores — secundaria", 5, min(20,len(db2e)), 10, key="sl_e2")
            pch(bar_h(db2e.head(top_se2), "RFCs","sub_cod","sub_nom","pct",
                      f"Top {top_se2} Subsectores Secundarios · Company {emp_sel}"))
        div()

    # Diversificación empresa
    sec("Diversificación — Actividades Secundarias")
    div_e  = get_div(df_e_a)
    solo_e = (div_e["act_sec"]==0).sum()
    con_e  = (div_e["act_sec"]>0).sum()
    prom_e = round(div_e["total_act"].mean(), 1)
    max_e  = div_e["total_act"].max()

    d1,d2,d3,d4 = st.columns(4)
    with d1: kpi("Solo principal",       f"{solo_e:,}", f"{round(solo_e/tr*100,1)}%")
    with d2: kpi("Con secundarias",      f"{con_e:,}",  f"{round(con_e/tr*100,1)}%")
    with d3: kpi("Promedio actividades", f"{prom_e}")
    with d4: kpi("Máx. actividades",     f"{max_e}")
    st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)

    ea3, eb3 = st.columns(2)
    with ea3:
        dist_e = (div_e["total_act"].value_counts()
                  .reset_index().rename(columns={"total_act":"N actividades","count":"RFCs"})
                  .sort_values("N actividades"))
        pch(grafica_dist(dist_e, tr, "Distribución: N° de actividades por RFC"))
    with eb3:
        top_div_e = (div_e[div_e["act_sec"]>0]
                     .sort_values("total_act", ascending=False).head(15)
                     [["rfc","company","tipo","total_act","act_sec"]]
                     .rename(columns={"rfc":"RFC","company":"Company","tipo":"Tipo",
                                      "total_act":"Total Act.","act_sec":"Secundarias"}))
        if not top_div_e.empty:
            pch(grafica_top_div(top_div_e, "Top 15 RFCs más diversificados"))

    combos_e = get_combos(df_e, df_e_s)
    if not combos_e.empty:
        pch(grafica_combos(combos_e, "Combinaciones Principal → Secundaria más frecuentes"))
    div()

    # Comparativo
    if comparar and emp_cmp:
        sec(f"Comparativo — Company {emp_sel}  vs  Company {emp_cmp}")
        df_c2 = df_p[df_p["company"] == emp_cmp]

        def top15(df_in, lbl):
            r = (df_in.groupby(["sub_cod","sub_nom"]).size()
                 .reset_index(name="RFCs").sort_values("RFCs", ascending=False).head(15))
            r["company"] = lbl; return r

        df_both = pd.concat([top15(df_e, f"Company {emp_sel}"),
                             top15(df_c2, f"Company {emp_cmp}")])
        fig = px.bar(df_both, x="RFCs", y="sub_cod", color="company",
                     orientation="h", barmode="group",
                     labels={"sub_cod":"","RFCs":"RFCs","company":""},
                     title="Comparativo de Subsectores Principales",
                     color_discrete_sequence=["#2E75B6","#F59E0B"],
                     custom_data=["sub_nom"])
        fig.update_traces(
            hovertemplate="<b>%{y}</b><br>%{customdata[0]}<br>RFCs: %{x}<extra></extra>")
        lay(fig, xaxis=dict(title="RFCs",**GRID),
            yaxis=dict(type="category", categoryorder="total ascending",
                       gridcolor="#2A3441"))
        pch(fig)

        tc2 = df_c2["rfc"].nunique()
        st.markdown("**Métricas comparadas**")
        st.dataframe(pd.DataFrame({
            "Métrica": ["Total RFCs","Sectores únicos",
                        "Subsectores únicos","Persona Moral","Persona Física"],
            f"Company {emp_sel}": [tr, ts, tb, nm, nf],
            f"Company {emp_cmp}": [tc2, df_c2["sec_cod"].nunique(),
                                   df_c2["sub_cod"].nunique(),
                                   (df_c2["tipo"]=="Persona Moral").sum(),
                                   (df_c2["tipo"]=="Persona Física").sum()],
        }), width="stretch", hide_index=True)
        div()

    # Tabla detalle
    with st.expander(f"📋  Ver detalle completo de RFCs — Company {emp_sel}"):
        st.dataframe(
            df_e_a[["rfc","tipo","tipo_act","sec_cod","sec_nom","sub_cod","sub_nom","pct"]]
            .rename(columns={"rfc":"RFC","tipo":"Tipo Persona","tipo_act":"Actividad",
                             "sec_cod":"Cód. Sector","sec_nom":"Sector",
                             "sub_cod":"Cód. Subsector","sub_nom":"Subsector",
                             "pct":"% Actividad"})
            .sort_values(["RFC","% Actividad"], ascending=[True,False]),
            width="stretch", hide_index=True)